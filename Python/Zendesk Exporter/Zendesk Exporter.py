"""
汎用Zendeskエクスポートツール v3.0
- 設定UIでブランド・カラム選択
- 認証情報は暗号化保存
- 質問/回答判定はRoleベース
- 除外タグ機能（オートコンプリート付き）
- メッセージングの回答数カウントをチケット単位で1に修正
- 統計集計の根拠をログファイルに出力
- システムメッセージを履歴には残すが統計からは除外
- チケット取得を最終更新日(updated)ベースに変更
- APIリクエストにリトライ処理を追加（502/503/429対応）
- チケット取得の日付バッファを前後40日に変更（Zendesk仕様に基づく）
- UI改善（v3.0）
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import pytz
import re
import os
import sys
import json
import logging
from cryptography.fernet import Fernet
from dateutil import parser
from typing import Dict, List, Set, Tuple, Any, Optional
import threading
import time
import glob

# アプリケーションのベースディレクトリを取得（exe対応）
def get_app_dir():
    """exeまたはスクリプトのあるディレクトリを取得"""
    if getattr(sys, 'frozen', False):
        # PyInstallerでexe化された場合
        return os.path.dirname(sys.executable)
    else:
        # 通常のPythonスクリプト
        return os.path.dirname(os.path.abspath(__file__))

APP_DIR = get_app_dir()

# ログ設定
def setup_logging():
    """日付別ログファイルを設定し、7日より古いログを削除"""
    log_dir = os.path.join(APP_DIR, "logs")
    
    # logsフォルダ作成
    os.makedirs(log_dir, exist_ok=True)
    
    # 今日のログファイル
    today = datetime.now().strftime("%Y-%m-%d")
    log_file = os.path.join(log_dir, f"export_{today}.log")
    
    # 7日より古いログを削除
    for old_log in glob.glob(os.path.join(log_dir, "export_*.log")):
        try:
            filename = os.path.basename(old_log)
            date_str = filename.replace("export_", "").replace(".log", "")
            log_date = datetime.strptime(date_str, "%Y-%m-%d")
            if (datetime.now() - log_date).days > 7:
                os.remove(old_log)
        except (ValueError, OSError):
            pass
    
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger("zendesk_exporter")

logger = setup_logging()


# =============================================================================
# カラー定義
# =============================================================================

class Colors:
    """アプリケーションカラー"""
    BG = "#f5f7fa"              # 背景（薄いグレー）
    BG_WHITE = "#ffffff"        # カード背景
    PRIMARY = "#2563eb"         # メインカラー（青）
    PRIMARY_HOVER = "#1d4ed8"   # ホバー時
    TEXT = "#1f2937"            # テキスト
    TEXT_SECONDARY = "#6b7280"  # サブテキスト
    BORDER = "#e5e7eb"          # ボーダー
    SUCCESS = "#10b981"         # 成功
    ERROR = "#ef4444"           # エラー
    TAG_BG = "#dbeafe"          # タグ背景
    TAG_TEXT = "#1e40af"        # タグテキスト


# =============================================================================
# 認証情報管理
# =============================================================================

class CredentialManager:
    """認証情報の暗号化・保存・読み込みを管理"""
    
    def __init__(self):
        self.key_file = os.path.join(APP_DIR, "zendesk_key.key")
        self.cred_file = os.path.join(APP_DIR, "zendesk_credentials.enc")
        self.key = self._load_or_create_key()
        self.fernet = Fernet(self.key)
    
    def _load_or_create_key(self) -> bytes:
        if os.path.exists(self.key_file):
            with open(self.key_file, "rb") as f:
                return f.read()
        else:
            key = Fernet.generate_key()
            with open(self.key_file, "wb") as f:
                f.write(key)
            return key
    
    def save_credentials(self, subdomain: str, email: str, api_token: str):
        data = json.dumps({
            "subdomain": subdomain,
            "email": email,
            "api_token": api_token
        })
        encrypted = self.fernet.encrypt(data.encode())
        with open(self.cred_file, "wb") as f:
            f.write(encrypted)
        logger.info("認証情報を保存しました")
    
    def load_credentials(self) -> Optional[Dict[str, str]]:
        if not os.path.exists(self.cred_file):
            return None
        try:
            with open(self.cred_file, "rb") as f:
                encrypted = f.read()
            decrypted = self.fernet.decrypt(encrypted)
            return json.loads(decrypted.decode())
        except Exception as e:
            logger.error(f"認証情報の読み込みエラー: {e}")
            return None


# =============================================================================
# Zendesk API
# =============================================================================

class ZendeskAPI:
    """Zendesk APIとの通信"""
    
    def __init__(self, subdomain: str, email: str, api_token: str):
        self.base_url = f"https://{subdomain}.zendesk.com/api/v2/"
        self.auth = (f"{email}/token", api_token)
        self._user_cache: Dict[int, Dict] = {}
    
    def test_connection(self) -> bool:
        try:
            response = requests.get(f"{self.base_url}users/me.json", auth=self.auth)
            return response.status_code == 200
        except Exception:
            return False
    
    def get_brands(self) -> Dict[int, str]:
        response = self._make_request("brands.json")
        return {brand['id']: brand['name'] for brand in response['brands']}
    
    def get_ticket_forms(self) -> Dict[int, str]:
        try:
            response = self._make_request("ticket_forms.json")
            return {form['id']: form['name'] for form in response['ticket_forms']}
        except Exception:
            return {}

    def get_tags(self) -> List[str]:
        try:
            tags = []
            url = f"{self.base_url}tags.json"
            
            while url:
                response = requests.get(url, auth=self.auth)
                response.raise_for_status()
                data = response.json()
                tags.extend([tag['name'] for tag in data.get('tags', [])])
                
                url = data.get('next_page')
                if len(tags) >= 5000:
                    break
            
            return tags
        except Exception as e:
            logger.error(f"タグ取得エラー: {e}")
            return []
    
    def get_user(self, user_id: int) -> Dict:
        if user_id in self._user_cache:
            return self._user_cache[user_id]
        
        try:
            response = self._make_request(f"users/{user_id}.json")
            user = response['user']
            self._user_cache[user_id] = user
            return user
        except Exception:
            return {"name": "Unknown", "role": "end-user", "id": user_id}
    
    def get_tickets(self, start_date: str, end_date: str) -> List[Dict]:
        query = f"updated>={start_date} updated<={end_date} type:ticket"
        return self._get_paginated_results(f"search.json?query={query}")
    
    def get_ticket_comments(self, ticket_id: int) -> List[Dict]:
        response = self._make_request(f"tickets/{ticket_id}/comments.json")
        return response['comments']
    
    def get_ticket_audits(self, ticket_id: int) -> List[Dict]:
        response = self._make_request(f"tickets/{ticket_id}/audits.json")
        return response['audits']
    
    def _make_request(self, endpoint: str, retries: int = 3) -> Dict:
        url = f"{self.base_url}{endpoint}"
        
        for attempt in range(retries):
            try:
                response = requests.get(url, auth=self.auth)
                response.raise_for_status()
                return response.json()
            except requests.exceptions.HTTPError as e:
                if response.status_code in (502, 503, 429) and attempt < retries - 1:
                    wait_time = 2 ** attempt
                    logger.warning(f"API エラー {response.status_code}、{wait_time}秒後にリトライ ({attempt + 1}/{retries})")
                    time.sleep(wait_time)
                    continue
                raise
    
    def _get_paginated_results(self, endpoint: str, retries: int = 3) -> List[Dict]:
        url = f"{self.base_url}{endpoint}"
        results = []
        
        while url:
            for attempt in range(retries):
                try:
                    response = requests.get(url, auth=self.auth)
                    response.raise_for_status()
                    data = response.json()
                    break
                except requests.exceptions.HTTPError as e:
                    if response.status_code in (502, 503, 429) and attempt < retries - 1:
                        wait_time = 2 ** attempt
                        logger.warning(f"API エラー {response.status_code}、{wait_time}秒後にリトライ ({attempt + 1}/{retries})")
                        time.sleep(wait_time)
                        continue
                    raise
            
            results.extend(data['results'])
            
            if len(results) >= 1000:
                break
            
            url = data.get('next_page')
            if url:
                logger.info(f"取得中... {len(results)}件")
        
        return results


# =============================================================================
# データ処理
# =============================================================================

class DataProcessor:
    """データ処理"""
    
    def __init__(self, api: ZendeskAPI):
        self.api = api
        self.jst = pytz.timezone('Asia/Tokyo')
    
    def clean_text(self, text: str) -> str:
        if not text:
            return ""
        text = re.sub(r'<(?!br\s*\/?)[^>]+>', '', text)
        text = text.replace("&nbsp;", " ").replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
        text = text.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
        text = text.replace('\r\n', '\n')
        text = re.sub(r'(\n{3,})', '\n\n', text)
        return text.strip()
    
    def format_date(self, date_string: str) -> str:
        if not date_string:
            return ""
        dt = parser.parse(date_string)
        return dt.astimezone(self.jst).strftime("%Y/%m/%d %H:%M")
    
    def format_messaging_date(self, msg: Dict) -> str:
        if "original_message" in msg and "received" in msg["original_message"]:
            timestamp = msg["original_message"]["received"]
            dt = parser.parse(timestamp)
        else:
            dt = datetime.fromtimestamp(msg["raw_timestamp"]/1000, tz=pytz.UTC)
        return dt.astimezone(self.jst).strftime("%Y/%m/%d %H:%M")
    
    def is_staff(self, user: Dict) -> bool:
        role = user.get('role', 'end-user')
        return role in ('admin', 'agent')
    
    def extract_messaging_history(self, audit_events: List[Dict]) -> List[Dict]:
        messages = []
        for event in audit_events:
            if event.get("type") == "ChatStartedEvent" and "history" in event.get("value", {}):
                for msg in event["value"]["history"]:
                    if msg.get("type") == "ChatMessage":
                        if msg.get("original_message_type") in ["form", "form_response"]:
                            continue
                        
                        actor_type = msg.get("actor_type", "")
                        message = msg.get("message", "")
                        
                        if message:
                            is_staff = actor_type in ("agent", "system")
                            
                            if "original_message" in msg and "received" in msg["original_message"]:
                                created_at = msg["original_message"]["received"]
                            else:
                                created_at = datetime.fromtimestamp(
                                    msg["timestamp"]/1000, 
                                    tz=pytz.UTC
                                ).strftime("%Y-%m-%dT%H:%M:%S.%fZ")
                            
                            messages.append({
                                "type": "staff" if is_staff else "user",
                                "created_at": created_at,
                                "content": self.clean_text(message),
                                "author_id": msg.get("actor_id"),
                                "actor_type": actor_type,
                                "original_message": msg.get("original_message", {}),
                                "raw_timestamp": msg.get("timestamp"),
                                "is_messaging": True
                            })
        return messages
    
    def process_tickets(self, tickets: List[Dict], brands: Dict[int, str], 
                       forms: Dict[int, str], selected_brands: List[int],
                       start_date: datetime, end_date: datetime,
                       exclude_tags: List[str] = None,
                       progress_callback=None) -> Tuple[Dict[str, List[Dict]], Dict[str, List[Dict]], Dict]:
        exclude_tags = exclude_tags or []
        
        brand_data: Dict[str, List[Dict]] = {brands[bid]: [] for bid in selected_brands}
        ai_agent_data: Dict[str, List[Dict]] = {brands[bid]: [] for bid in selected_brands}
        
        stats_data = self._initialize_stats_data(start_date, end_date, 
                                                  [brands[bid] for bid in selected_brands])
        
        sorted_tickets = sorted(tickets, key=lambda t: t['created_at'])
        
        total = len(sorted_tickets)
        for idx, ticket in enumerate(sorted_tickets):
            if progress_callback and idx % 10 == 0:
                progress_callback(f"処理中... {idx}/{total}件")
            
            if ticket.get('status') == 'deleted':
                continue
            
            ticket_tags = ticket.get('tags', [])
            if any(tag in ticket_tags for tag in exclude_tags):
                continue

            brand_id = ticket.get('brand_id')
            if brand_id not in selected_brands:
                continue
            
            brand_name = brands[brand_id]
            is_ai_agent = ticket.get('support_type') == 'ai_agent'
            
            comments = self.api.get_ticket_comments(ticket['id'])
            
            audits = self.api.get_ticket_audits(ticket['id'])
            messaging_comments = []
            for audit in audits:
                messaging_comments.extend(self.extract_messaging_history(audit['events']))
            
            combined_comments = []
            
            for comment in comments:
                author_id = comment.get('author_id')
                author = self.api.get_user(author_id)
                is_staff_comment = self.is_staff(author)
                is_public = comment.get('public', True)
                
                if not is_public and is_staff_comment:
                    continue
                
                comment_date = parser.parse(comment['created_at']).astimezone(self.jst)
                
                if not is_ai_agent and is_public:
                    self._update_stats(stats_data, comment_date, brand_name, 
                                       author_id, is_staff_comment, ticket['id'], "Email")
                
                if start_date <= comment_date <= end_date:
                    combined_comments.append({
                        "type": "staff" if is_staff_comment else "user",
                        "created_at": comment['created_at'],
                        "content": self.clean_text(comment.get('html_body', '')),
                        "author_id": author_id,
                        "author_name": author.get('name', 'Unknown'),
                        "is_messaging": False
                    })
            
            messaging_staff_reply_counted = False

            for msg in messaging_comments:
                msg_date = parser.parse(msg['created_at']).astimezone(self.jst)
                
                if not is_ai_agent:
                    author_id = msg.get('author_id')
                    is_staff_msg = msg['type'] == 'staff'
                    actor_type = msg.get('actor_type')
                    
                    if is_staff_msg:
                        if actor_type == 'agent':
                            if not messaging_staff_reply_counted:
                                self._update_stats(stats_data, msg_date, brand_name, 
                                                   author_id, is_staff_msg, ticket['id'], "Messaging(1st)")
                                messaging_staff_reply_counted = True
                    else:
                        self._update_stats(stats_data, msg_date, brand_name, 
                                           author_id, is_staff_msg, ticket['id'], "Messaging(User)")
                
                if start_date <= msg_date <= end_date:
                    author_id = msg.get('author_id')
                    if author_id:
                        author = self.api.get_user(author_id)
                        msg['author_name'] = author.get('name', 'Unknown')
                    else:
                        msg['author_name'] = 'Unknown'
                    combined_comments.append(msg)
            
            if not combined_comments:
                continue
            
            combined_comments = sorted(combined_comments, key=lambda c: c['created_at'])
            
            requester = self.api.get_user(ticket['requester_id'])
            form_name = forms.get(ticket.get('ticket_form_id'), '')
            
            ticket_rows = []
            for comment in combined_comments:
                comment_type = "回答" if comment['type'] == 'staff' else "質問"
                
                if comment.get('is_messaging'):
                    comment_date_str = self.format_messaging_date(comment)
                else:
                    comment_date_str = self.format_date(comment['created_at'])
                
                row = {
                    'ticket_id': ticket['id'],
                    'requester_name': requester.get('name', ''),
                    'brand_name': brand_name,
                    'form_name': form_name,
                    'status': ticket.get('status', ''),
                    'ticket_created_at': self.format_date(ticket['created_at']),
                    'comment_type': comment_type,
                    'comment_author': comment.get('author_name', 'Unknown'),
                    'comment_date': comment_date_str,
                    'comment_body': comment.get('content', ''),
                    'tags': ', '.join(ticket.get('tags', [])),
                    'author_id': comment.get('author_id'),
                }
                ticket_rows.append(row)
            
            if ticket_rows:
                ticket_data = {
                    'ticket_id': ticket['id'],
                    'rows': ticket_rows
                }
                
                if is_ai_agent:
                    ai_agent_data[brand_name].append(ticket_data)
                else:
                    brand_data[brand_name].append(ticket_data)
        
        return brand_data, ai_agent_data, stats_data
    
    def _initialize_stats_data(self, start_date: datetime, end_date: datetime, 
                               brand_names: List[str]) -> Dict:
        stats_data = {}
        current = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end = end_date.replace(hour=23, minute=59, second=59)
        
        while current <= end:
            date_str = current.strftime("%Y/%m/%d")
            stats_data[date_str] = {}
            for brand in brand_names:
                stats_data[date_str][f"{brand}_UU"] = set()
                stats_data[date_str][f"{brand}_回答数"] = 0
            current += timedelta(days=1)
        
        return stats_data
    
    def _update_stats(self, stats_data: Dict, comment_date: datetime, 
                     brand_name: str, author_id: int, is_staff: bool,
                     ticket_id: int = None, source: str = "Unknown"):
        date_str = comment_date.strftime("%Y/%m/%d")
        if date_str not in stats_data:
            return
        
        if is_staff:
            stats_data[date_str][f"{brand_name}_回答数"] += 1
            logger.info(f"[統計:回答+1] 日付:{date_str}, ブランド:{brand_name}, チケットID:{ticket_id}, ソース:{source}")
        else:
            if author_id:
                if author_id not in stats_data[date_str][f"{brand_name}_UU"]:
                    stats_data[date_str][f"{brand_name}_UU"].add(author_id)
                    logger.info(f"[統計:UU+1] 日付:{date_str}, ブランド:{brand_name}, チケットID:{ticket_id}, ユーザーID:{author_id}, ソース:{source}")


# =============================================================================
# Excel出力
# =============================================================================

AVAILABLE_COLUMNS = [
    ('ticket_id', 'チケットID'),
    ('requester_name', '依頼者名'),
    ('brand_name', 'ブランド'),
    ('form_name', 'フォーム'),
    ('status', 'ステータス'),
    ('ticket_created_at', 'チケット作成日時'),
    ('comment_type', '種別'),
    ('comment_author', '投稿者'),
    ('comment_date', 'コメント日時'),
    ('comment_body', 'コメント内容'),
    ('tags', 'タグ'),
]


class ExcelExporter:
    """Excel出力"""
    
    def __init__(self, selected_columns: List[str]):
        self.selected_columns = selected_columns
        self.workbook = Workbook()
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.separator_fill = PatternFill(start_color="FFFFD1", end_color="FFFFD1", fill_type="solid")
        self.question_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_brand_sheet(self, brand_name: str, tickets_data: List[Dict], 
                          sheet_suffix: str = ""):
        sheet_name = f"{brand_name}{sheet_suffix}"[:31]
        sheet = self.workbook.create_sheet(title=sheet_name)
        
        headers = [col[1] for col in AVAILABLE_COLUMNS if col[0] in self.selected_columns]
        sheet.append(headers)
        
        for cell in sheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
        
        for ticket_data in tickets_data:
            separator_row = [''] * len(headers)
            separator_row[0] = f"--- チケット {ticket_data['ticket_id']} ---"
            sheet.append(separator_row)
            for cell in sheet[sheet.max_row]:
                cell.fill = self.separator_fill
            
            for row_data in ticket_data['rows']:
                row = []
                for col_id, _ in AVAILABLE_COLUMNS:
                    if col_id in self.selected_columns:
                        row.append(row_data.get(col_id, ''))
                sheet.append(row)
                
                if row_data.get('comment_type') == '質問':
                    for cell in sheet[sheet.max_row]:
                        cell.fill = self.question_fill
        
        self._apply_styles(sheet)
    
    def create_stats_sheet(self, stats_data: Dict, brand_names: List[str]):
        sheet = self.workbook.create_sheet(title="統計")
        
        headers = ["日付"]
        for brand in brand_names:
            headers.extend([f"{brand}_UU", f"{brand}_回答数"])
        headers.extend(["合計_UU", "合計_回答数"])
        sheet.append(headers)
        
        for cell in sheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
        
        for date_str in sorted(stats_data.keys()):
            day_stats = stats_data[date_str]
            row = [date_str]
            
            total_uu = set()
            total_answers = 0
            
            for brand in brand_names:
                brand_uu = day_stats.get(f"{brand}_UU", set())
                brand_answers = day_stats.get(f"{brand}_回答数", 0)
                
                row.extend([len(brand_uu), brand_answers])
                total_uu.update(brand_uu)
                total_answers += brand_answers
            
            row.extend([len(total_uu), total_answers])
            sheet.append(row)
        
        self._apply_styles(sheet)
    
    def _apply_styles(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = self.border
        
        for i, col in enumerate(sheet.columns, 1):
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[get_column_letter(i)].width = adjusted_width
        
        sheet.freeze_panes = "A2"
    
    def save(self, output_path: str):
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        self.workbook.save(output_path)
        logger.info(f"保存完了: {output_path}")


# =============================================================================
# 共通エクスポート処理
# =============================================================================

def move_old_exports_to_archive(output_folder: str):
    """14日以上前のエクスポートファイルをoldフォルダに移動"""
    import shutil
    
    old_folder = os.path.join(output_folder, "old")
    today = datetime.now()
    
    for filepath in glob.glob(os.path.join(output_folder, "zendesk_export_*.xlsx")):
        filename = os.path.basename(filepath)
        # ファイル名から終了日を取得: zendesk_export_2025-01-13_to_2025-01-19.xlsx
        try:
            # _to_の後の日付を取得
            match = re.search(r'_to_(\d{4}-\d{2}-\d{2})\.xlsx$', filename)
            if match:
                end_date_str = match.group(1)
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
                
                if (today - end_date).days >= 14:
                    # oldフォルダ作成
                    os.makedirs(old_folder, exist_ok=True)
                    # 移動
                    dest_path = os.path.join(old_folder, filename)
                    shutil.move(filepath, dest_path)
                    logger.info(f"古いファイルを移動: {filename} -> old/")
        except (ValueError, OSError) as e:
            logger.warning(f"ファイル移動エラー: {filename} - {e}")


def execute_export(
    api: ZendeskAPI,
    brands: Dict[int, str],
    forms: Dict[int, str],
    selected_brand_ids: List[int],
    selected_columns: List[str],
    exclude_tags: List[str],
    start_date: datetime,
    end_date: datetime,
    ai_agent_separate: bool,
    output_path: str,
    progress_callback=None
):
    """
    エクスポート処理の共通関数
    
    Args:
        api: ZendeskAPI インスタンス
        brands: ブランドID→名前のマッピング
        forms: フォームID→名前のマッピング
        selected_brand_ids: 選択されたブランドIDリスト
        selected_columns: 選択されたカラムIDリスト
        exclude_tags: 除外タグリスト
        start_date: 開始日時
        end_date: 終了日時
        ai_agent_separate: AIエージェントを別シートにするか
        output_path: 出力ファイルパス
        progress_callback: 進捗コールバック関数(msg: str) -> None
    """
    def progress(msg: str):
        if progress_callback:
            progress_callback(msg)
    
    # 古いファイルをoldフォルダに移動
    output_folder = os.path.dirname(output_path)
    if output_folder:
        move_old_exports_to_archive(output_folder)
    
    progress("チケット取得中...")
    
    processor = DataProcessor(api)
    
    buffer_start = (start_date - timedelta(days=40)).strftime("%Y-%m-%d")
    buffer_end = (end_date + timedelta(days=40)).strftime("%Y-%m-%d")
    
    tickets = api.get_tickets(buffer_start, buffer_end)
    progress(f"{len(tickets)}件のチケットを取得しました。処理中...")
    
    brand_data, ai_agent_data, stats_data = processor.process_tickets(
        tickets, brands, forms,
        selected_brand_ids, start_date, end_date,
        exclude_tags=exclude_tags,
        progress_callback=progress_callback
    )
    
    progress("Excel出力中...")
    
    exporter = ExcelExporter(selected_columns)
    brand_names = [brands[bid] for bid in selected_brand_ids]
    
    for brand_name, tickets_data in brand_data.items():
        if tickets_data:
            exporter.create_brand_sheet(brand_name, tickets_data)
    
    if ai_agent_separate:
        for brand_name, tickets_data in ai_agent_data.items():
            if tickets_data:
                exporter.create_brand_sheet(brand_name, tickets_data, sheet_suffix="_AI")
    
    exporter.create_stats_sheet(stats_data, brand_names)
    exporter.save(output_path)
    
    progress("完了")


# =============================================================================
# UIコンポーネント
# =============================================================================

class TagSelector(tk.Frame):
    """タグ選択ウィジェット（サジェスト式ドロップダウン）"""
    
    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=Colors.BG_WHITE, **kwargs)
        self.all_tags: List[str] = []
        self.selected_tags: List[str] = []
        self.enabled = False
        self._dropdown_window: Optional[tk.Toplevel] = None
        self._selected_index = -1
        self._init_ui()
    
    def _init_ui(self):
        # 検索バー
        search_frame = tk.Frame(self, bg=Colors.BG_WHITE)
        search_frame.pack(fill=tk.X, pady=(0, 12))
        
        tk.Label(search_frame, text="🔍", bg=Colors.BG_WHITE, 
                font=("Meiryo UI", 11)).pack(side=tk.LEFT)
        
        self.tag_entry = tk.Entry(search_frame, width=45,
                                  font=("Meiryo UI", 10), relief="solid", bd=1)
        self.tag_entry.pack(side=tk.LEFT, padx=(8, 0), ipady=6)
        
        self.tag_entry.bind('<FocusIn>', self._on_focus_in)
        self.tag_entry.bind('<FocusOut>', self._on_focus_out)
        self.tag_entry.bind('<KeyRelease>', self._on_key_release)
        self.tag_entry.bind('<Return>', self._add_highlighted_tag)
        self.tag_entry.bind('<Down>', self._move_selection_down)
        self.tag_entry.bind('<Up>', self._move_selection_up)
        self.tag_entry.bind('<Escape>', self._hide_dropdown)
        
        # 選択済みタグ表示
        tk.Label(self, text="除外するタグ:", bg=Colors.BG_WHITE,
                fg=Colors.TEXT_SECONDARY, font=("Meiryo UI", 9)).pack(anchor=tk.W, pady=(4, 0))
        
        self.selected_frame = tk.Frame(self, bg=Colors.BG_WHITE)
        self.selected_frame.pack(fill=tk.X, pady=(4, 0))
        
        self.no_tags_label = tk.Label(self.selected_frame, text="（なし）",
                                      bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                                      font=("Meiryo UI", 9))
        self.no_tags_label.pack(anchor=tk.W)
    
    def _on_focus_in(self, event):
        """フォーカス時"""
        if self.enabled:
            self.after(50, self._show_dropdown)
    
    def _on_focus_out(self, event):
        """フォーカスアウト時"""
        self.after(200, self._delayed_focus_check)
    
    def _delayed_focus_check(self):
        """遅延してフォーカスをチェック"""
        try:
            focused = self.focus_get()
            if self._dropdown_window and focused:
                widget_str = str(focused)
                dropdown_str = str(self._dropdown_window)
                if dropdown_str in widget_str or widget_str == dropdown_str:
                    return
        except:
            pass
        
        self._hide_dropdown()
    
    def _on_key_release(self, event):
        """キー入力時にサジェスト更新"""
        if event.keysym in ('Up', 'Down', 'Return', 'Escape', 'Shift_L', 'Shift_R', 'Control_L', 'Control_R'):
            return
        
        if self.enabled:
            if not self._dropdown_window:
                self._show_dropdown()
            else:
                self._update_suggestions()
    
    def _show_dropdown(self):
        """サジェストドロップダウンを表示"""
        if self._dropdown_window:
            return
        
        if not self.enabled:
            return
        
        # エントリーの位置を取得
        self.tag_entry.update_idletasks()
        x = self.tag_entry.winfo_rootx()
        y = self.tag_entry.winfo_rooty() + self.tag_entry.winfo_height()
        width = self.tag_entry.winfo_width()
        
        # ドロップダウンウィンドウ作成
        self._dropdown_window = tk.Toplevel(self)
        self._dropdown_window.wm_overrideredirect(True)
        self._dropdown_window.wm_geometry(f"{width}x200+{x}+{y}")
        
        # リストボックス
        frame = tk.Frame(self._dropdown_window, bg=Colors.BORDER)
        frame.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        
        self.listbox = tk.Listbox(frame, font=("Meiryo UI", 10),
                                  relief="flat", bd=0, 
                                  selectbackground=Colors.PRIMARY,
                                  selectforeground="white",
                                  activestyle="none", 
                                  exportselection=False,
                                  highlightthickness=0)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = tk.Scrollbar(frame, orient=tk.VERTICAL, command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=scrollbar.set)
        
        self.listbox.bind('<Double-Button-1>', self._on_listbox_double_click)
        
        # マウスホイールイベントを自前で処理（親への伝播を防ぐ）
        self.listbox.bind('<MouseWheel>', self._on_listbox_scroll)
        self.listbox.bind('<Button-4>', self._on_listbox_scroll)  # Linux用
        self.listbox.bind('<Button-5>', self._on_listbox_scroll)  # Linux用
        
        self._selected_index = -1
        self._update_suggestions()
        
        # 位置追従開始
        self._track_position()
    
    def _track_position(self):
        """ドロップダウンの位置を追従"""
        if not self._dropdown_window:
            return
        try:
            x = self.tag_entry.winfo_rootx()
            y = self.tag_entry.winfo_rooty() + self.tag_entry.winfo_height()
            width = self.tag_entry.winfo_width()
            self._dropdown_window.wm_geometry(f"{width}x200+{x}+{y}")
            self.after(50, self._track_position)
        except:
            pass
    
    def _hide_dropdown(self, event=None):
        """ドロップダウンを非表示"""
        if self._dropdown_window:
            self._dropdown_window.destroy()
            self._dropdown_window = None
            self._selected_index = -1
    
    def _on_listbox_scroll(self, event):
        """リストボックスのスクロール（親への伝播を防ぐ）"""
        if event.num == 4:  # Linux scroll up
            self.listbox.yview_scroll(-3, "units")
        elif event.num == 5:  # Linux scroll down
            self.listbox.yview_scroll(3, "units")
        else:  # Windows/Mac
            self.listbox.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"
    
    def _update_suggestions(self):
        """サジェストリストを更新"""
        if not self._dropdown_window or not hasattr(self, 'listbox') or not self.listbox:
            return
        
        self.listbox.delete(0, tk.END)
        
        search_text = self.tag_entry.get().lower().strip()
        
        # フィルタリング
        if not search_text:
            suggestions = [t for t in self.all_tags if t not in self.selected_tags]
        else:
            suggestions = [t for t in self.all_tags 
                          if search_text in t.lower() and t not in self.selected_tags]
        
        for tag in suggestions[:100]:
            self.listbox.insert(tk.END, tag)
        
        # 選択をリセット
        self._selected_index = -1
        if self.listbox.size() > 0:
            self._selected_index = 0
            self.listbox.selection_set(0)
            self.listbox.see(0)
    
    def _move_selection_down(self, event):
        """選択を下に移動"""
        if not self._dropdown_window or not self.listbox:
            if self.enabled:
                self._show_dropdown()
            return "break"
        
        if self.listbox.size() == 0:
            return "break"
        
        self._selected_index = min(self._selected_index + 1, self.listbox.size() - 1)
        self.listbox.selection_clear(0, tk.END)
        self.listbox.selection_set(self._selected_index)
        self.listbox.see(self._selected_index)
        return "break"
    
    def _move_selection_up(self, event):
        """選択を上に移動"""
        if not self._dropdown_window or not self.listbox:
            return "break"
        
        if self.listbox.size() == 0:
            return "break"
        
        self._selected_index = max(self._selected_index - 1, 0)
        self.listbox.selection_clear(0, tk.END)
        self.listbox.selection_set(self._selected_index)
        self.listbox.see(self._selected_index)
        return "break"
    
    def _on_listbox_scroll(self, event):
        """リストボックスのスクロール（親への伝播を防ぐ）"""
        if event.num == 4:  # Linux scroll up
            self.listbox.yview_scroll(-3, "units")
        elif event.num == 5:  # Linux scroll down
            self.listbox.yview_scroll(3, "units")
        else:  # Windows/Mac
            self.listbox.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"
    
    def _on_listbox_double_click(self, event):
        """リストボックスダブルクリック時"""
        selection = self.listbox.curselection()
        if selection:
            tag = self.listbox.get(selection[0])
            self._add_tag(tag)
        self._hide_dropdown()
        # 明示的にフォーカスを外す
        self.after(10, lambda: self.winfo_toplevel().focus_set())
        return "break"
    
    def _add_highlighted_tag(self, event=None):
        """ハイライトされたタグを追加"""
        if not self._dropdown_window or not self.listbox:
            return
        
        selection = self.listbox.curselection()
        if selection:
            tag = self.listbox.get(selection[0])
            self._add_tag(tag)
    
    def _add_tag(self, tag: str):
        """タグを追加"""
        if tag and tag not in self.selected_tags and tag in self.all_tags:
            self.selected_tags.append(tag)
            self._update_selected_display()
            self.tag_entry.delete(0, tk.END)
            self._update_suggestions()
    
    def set_tags(self, tags: List[str]):
        self.all_tags = sorted(tags)
        self.enabled = True
    
    def get_selected_tags(self) -> List[str]:
        return self.selected_tags.copy()
    
    def _remove_tag(self, tag: str):
        if tag in self.selected_tags:
            self.selected_tags.remove(tag)
            self._update_selected_display()
            if self._dropdown_window:
                self._update_suggestions()
    
    def _update_selected_display(self):
        for widget in self.selected_frame.winfo_children():
            widget.destroy()
        
        if not self.selected_tags:
            tk.Label(self.selected_frame, text="（なし）", bg=Colors.BG_WHITE,
                    fg=Colors.TEXT_SECONDARY, font=("Meiryo UI", 9)).pack(anchor=tk.W)
            return
        
        row_frame = None
        for i, tag in enumerate(self.selected_tags):
            if i % 4 == 0:
                row_frame = tk.Frame(self.selected_frame, bg=Colors.BG_WHITE)
                row_frame.pack(fill=tk.X, pady=2)
            
            tag_frame = tk.Frame(row_frame, bg=Colors.TAG_BG, padx=8, pady=2)
            tag_frame.pack(side=tk.LEFT, padx=(0, 8))
            
            tk.Label(tag_frame, text=tag, bg=Colors.TAG_BG, fg=Colors.TAG_TEXT,
                    font=("Meiryo UI", 9)).pack(side=tk.LEFT)
            
            remove_btn = tk.Label(tag_frame, text=" ✕", bg=Colors.TAG_BG, fg=Colors.TAG_TEXT,
                                 font=("Meiryo UI", 9), cursor="hand2")
            remove_btn.pack(side=tk.LEFT)
            remove_btn.bind('<Button-1>', lambda e, t=tag: self._remove_tag(t))


class Card(tk.Frame):
    """カードコンポーネント"""
    
    def __init__(self, parent, title: str, **kwargs):
        super().__init__(parent, bg=Colors.BG_WHITE, padx=20, pady=16, **kwargs)
        
        # タイトル
        self.title_label = tk.Label(self, text=title, bg=Colors.BG_WHITE, fg=Colors.TEXT,
                              font=("Meiryo UI", 11, "bold"))
        self.title_label.pack(anchor=tk.W, pady=(0, 12))
        
        # コンテンツフレーム
        self.content = tk.Frame(self, bg=Colors.BG_WHITE)
        self.content.pack(fill=tk.BOTH, expand=True)
        
        # オーバーレイ（初期は非表示）
        self._overlay = None
    
    def set_disabled(self, disabled: bool):
        """カードの無効化状態を設定"""
        if disabled:
            self._show_overlay()
        else:
            self._hide_overlay()
    
    def _show_overlay(self):
        """半透明オーバーレイを表示"""
        if self._overlay:
            return
        
        self._overlay = tk.Frame(self, bg='#e0e0e0')
        self._overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        
        # オーバーレイ上のクリックを無効化
        self._overlay.bind('<Button-1>', lambda e: 'break')
        self._overlay.bind('<ButtonRelease-1>', lambda e: 'break')
        
        # 透過風の見た目にするためタイトルだけ表示
        if self.title_label.cget('text'):
            overlay_title = tk.Label(self._overlay, text=self.title_label.cget('text'),
                                    bg='#e0e0e0', fg='#999999',
                                    font=("Meiryo UI", 11, "bold"))
            overlay_title.pack(anchor=tk.W, padx=20, pady=(16, 0))
    
    def _hide_overlay(self):
        """オーバーレイを非表示"""
        if self._overlay:
            self._overlay.destroy()
            self._overlay = None


# =============================================================================
# メインUI
# =============================================================================

class SettingsUI:
    """設定UI"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Zendesk Exporter")
        self.root.geometry("800x900")
        self.root.configure(bg=Colors.BG)
        
        # フォント設定
        default_font = ("Meiryo UI", 10)
        self.root.option_add("*Font", default_font)
        
        self.cred_manager = CredentialManager()
        self.api: Optional[ZendeskAPI] = None
        self.brands: Dict[int, str] = {}
        self.forms: Dict[int, str] = {}
        self.tags: List[str] = []
        
        self.brand_vars: Dict[int, tk.BooleanVar] = {}
        self.column_vars: Dict[str, tk.BooleanVar] = {}
        
        self._init_ui()
        self._load_saved_credentials()
        
        # 入力欄以外をクリックしたらフォーカスを外す
        self.root.bind_all('<Button-1>', self._on_global_click, add='+')
    
    def _on_global_click(self, event):
        """グローバルクリックハンドラ"""
        widget = event.widget
        # EntryとListbox以外をクリックしたらフォーカスをルートに
        if not isinstance(widget, (tk.Entry, tk.Listbox)):
            self.root.focus_set()
    
    def _init_ui(self):
        # ヘッダー
        header = tk.Frame(self.root, bg=Colors.PRIMARY, height=56)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(header, text="📊 Zendesk Exporter", bg=Colors.PRIMARY, fg="white",
                font=("Meiryo UI", 14, "bold")).pack(side=tk.LEFT, padx=20, pady=12)
        
        tk.Label(header, text="v3.0", bg=Colors.PRIMARY, fg="white",
                font=("Meiryo UI", 10)).pack(side=tk.RIGHT, padx=20, pady=14)
        
        # 自動実行設定保存ボタン
        self.save_auto_btn = tk.Button(header, text="🤖 自動実行用に保存", 
                                       command=self._show_auto_config_dialog,
                                       bg="#10b981", fg="white", font=("Meiryo UI", 9, "bold"),
                                       relief="flat", padx=12, pady=4, cursor="hand2")
        self.save_auto_btn.pack(side=tk.RIGHT, padx=(0, 12), pady=12)
        
        # スクロール可能なコンテンツ
        canvas = tk.Canvas(self.root, bg=Colors.BG, highlightthickness=0)
        scrollbar = tk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=Colors.BG)
        
        scrollable_frame.bind("<Configure>", 
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        # キャンバスの幅に合わせてフレームの幅を調整
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw", tags="frame"),
            width=e.width
        ))
        canvas.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        
        # メインコンテンツ
        main_frame = tk.Frame(scrollable_frame, bg=Colors.BG)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # === 認証情報カード ===
        auth_card = Card(main_frame, "🔐 Zendesk 認証情報")
        auth_card.pack(fill=tk.X, pady=(0, 16))
        
        # サブドメイン
        row1 = tk.Frame(auth_card.content, bg=Colors.BG_WHITE)
        row1.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row1, text="サブドメイン", bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                font=("Meiryo UI", 9), width=12, anchor="w").pack(side=tk.LEFT)
        self.subdomain_entry = tk.Entry(row1, font=("Meiryo UI", 10), relief="solid", bd=1, width=30)
        self.subdomain_entry.pack(side=tk.LEFT, ipady=4)
        tk.Label(row1, text=".zendesk.com", bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                font=("Meiryo UI", 9)).pack(side=tk.LEFT, padx=(4, 0))
        
        # メール
        row2 = tk.Frame(auth_card.content, bg=Colors.BG_WHITE)
        row2.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row2, text="メール", bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                font=("Meiryo UI", 9), width=12, anchor="w").pack(side=tk.LEFT)
        self.email_entry = tk.Entry(row2, font=("Meiryo UI", 10), relief="solid", bd=1, width=35)
        self.email_entry.pack(side=tk.LEFT, ipady=4)
        
        # APIトークン
        row3 = tk.Frame(auth_card.content, bg=Colors.BG_WHITE)
        row3.pack(fill=tk.X, pady=(0, 12))
        tk.Label(row3, text="APIトークン", bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                font=("Meiryo UI", 9), width=12, anchor="w").pack(side=tk.LEFT)
        self.token_entry = tk.Entry(row3, font=("Meiryo UI", 10), relief="solid", bd=1, 
                                   width=35, show="•")
        self.token_entry.pack(side=tk.LEFT, ipady=4)
        
        # ボタン
        btn_frame = tk.Frame(auth_card.content, bg=Colors.BG_WHITE)
        btn_frame.pack(fill=tk.X, pady=(0, 8))
        
        connect_btn = tk.Button(btn_frame, text="接続", command=self._connect,
                               bg=Colors.PRIMARY, fg="white", font=("Meiryo UI", 10, "bold"),
                               relief="flat", padx=20, pady=6, cursor="hand2")
        connect_btn.pack(side=tk.LEFT)
        
        save_btn = tk.Button(btn_frame, text="認証情報を保存", command=self._save_credentials,
                            bg=Colors.BG, fg=Colors.TEXT, font=("Meiryo UI", 10),
                            relief="solid", bd=1, padx=12, pady=6, cursor="hand2")
        save_btn.pack(side=tk.LEFT, padx=(12, 0))
        
        self.status_label = tk.Label(auth_card.content, text="⚠️ まずZendeskに接続してください", bg=Colors.BG_WHITE,
                                    fg=Colors.PRIMARY, font=("Meiryo UI", 10, "bold"))
        self.status_label.pack(anchor=tk.W)
        
        # === ブランド選択カード ===
        self.brand_card = Card(main_frame, "🏷️ ブランド選択")
        self.brand_card.pack(fill=tk.X, pady=(0, 16))
        
        self.brand_frame = tk.Frame(self.brand_card.content, bg=Colors.BG_WHITE)
        self.brand_frame.pack(fill=tk.X)
        
        tk.Label(self.brand_frame, text="接続後にブランドが表示されます", bg=Colors.BG_WHITE,
                fg=Colors.TEXT_SECONDARY, font=("Meiryo UI", 9)).pack(anchor=tk.W)
        
        # === 除外タグカード ===
        self.tag_card = Card(main_frame, "🚫 除外タグ設定")
        self.tag_card.pack(fill=tk.X, pady=(0, 16))
        
        self.tag_selector = TagSelector(self.tag_card.content)
        self.tag_selector.pack(fill=tk.X)
        
        # === カラム選択カード ===
        self.column_card = Card(main_frame, "📋 出力カラム選択")
        self.column_card.pack(fill=tk.X, pady=(0, 16))
        
        column_grid = tk.Frame(self.column_card.content, bg=Colors.BG_WHITE)
        column_grid.pack(fill=tk.X)
        
        for i, (col_id, col_name) in enumerate(AVAILABLE_COLUMNS):
            var = tk.BooleanVar(value=True)
            self.column_vars[col_id] = var
            
            # チケットIDは必須（選択不可）
            if col_id == 'ticket_id':
                cb = tk.Checkbutton(column_grid, text=f"{col_name}（必須）", variable=var,
                                   bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY, font=("Meiryo UI", 9),
                                   activebackground=Colors.BG_WHITE, selectcolor=Colors.BG_WHITE,
                                   state='disabled')
            else:
                cb = tk.Checkbutton(column_grid, text=col_name, variable=var,
                                   bg=Colors.BG_WHITE, fg=Colors.TEXT, font=("Meiryo UI", 9),
                                   activebackground=Colors.BG_WHITE, selectcolor=Colors.BG_WHITE,
                                   command=self._update_preview)
            cb.grid(row=i // 4, column=i % 4, sticky=tk.W, padx=(0, 16), pady=2)
        
        # === 期間設定カード ===
        self.date_card = Card(main_frame, "📅 取得期間")
        self.date_card.pack(fill=tk.X, pady=(0, 16))
        
        date_frame = tk.Frame(self.date_card.content, bg=Colors.BG_WHITE)
        date_frame.pack(fill=tk.X)
        
        tk.Label(date_frame, text="開始日", bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                font=("Meiryo UI", 9)).pack(side=tk.LEFT)
        self.start_date_entry = tk.Entry(date_frame, font=("Meiryo UI", 10), 
                                        relief="solid", bd=1, width=12)
        self.start_date_entry.pack(side=tk.LEFT, padx=(8, 0), ipady=4)
        
        tk.Label(date_frame, text="〜", bg=Colors.BG_WHITE, fg=Colors.TEXT,
                font=("Meiryo UI", 10)).pack(side=tk.LEFT, padx=12)
        
        tk.Label(date_frame, text="終了日", bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                font=("Meiryo UI", 9)).pack(side=tk.LEFT)
        self.end_date_entry = tk.Entry(date_frame, font=("Meiryo UI", 10),
                                      relief="solid", bd=1, width=12)
        self.end_date_entry.pack(side=tk.LEFT, padx=(8, 0), ipady=4)
        
        tk.Label(date_frame, text="（形式: YYYY-MM-DD）", bg=Colors.BG_WHITE,
                fg=Colors.TEXT_SECONDARY, font=("Meiryo UI", 9)).pack(side=tk.LEFT, padx=(16, 0))
        
        # デフォルト日付
        today = datetime.now()
        start_of_week = today - timedelta(days=today.weekday() + 7)
        end_of_week = start_of_week + timedelta(days=6)
        self.start_date_entry.insert(0, start_of_week.strftime("%Y-%m-%d"))
        self.end_date_entry.insert(0, end_of_week.strftime("%Y-%m-%d"))
        
        # === オプションカード ===
        self.option_card = Card(main_frame, "⚙️ オプション")
        self.option_card.pack(fill=tk.X, pady=(0, 16))
        
        self.ai_agent_separate = tk.BooleanVar(value=True)
        self.ai_agent_cb = tk.Checkbutton(self.option_card.content, text="AIエージェントチケットを別シートに出力",
                           variable=self.ai_agent_separate, bg=Colors.BG_WHITE, fg=Colors.TEXT,
                           font=("Meiryo UI", 10), activebackground=Colors.BG_WHITE,
                           selectcolor=Colors.BG_WHITE, command=self._update_preview)
        self.ai_agent_cb.pack(anchor=tk.W)
        
        # === プレビューカード ===
        self.preview_card = Card(main_frame, "👁️ 出力イメージ")
        self.preview_card.pack(fill=tk.X, pady=(0, 16))
        
        self.preview_frame = tk.Frame(self.preview_card.content, bg=Colors.BG_WHITE)
        self.preview_frame.pack(fill=tk.X)
        
        # === 実行ボタン ===
        self.exec_card = Card(main_frame, "")
        self.exec_card.pack(fill=tk.X, pady=(0, 20))
        
        self.export_btn = tk.Button(self.exec_card.content, text="📥 エクスポート実行", 
                              command=self._execute_export,
                              bg=Colors.PRIMARY, fg="white", font=("Meiryo UI", 12, "bold"),
                              relief="flat", padx=32, pady=12, cursor="hand2")
        self.export_btn.pack(fill=tk.X)
        
        # 進捗表示
        self.progress_label = tk.Label(self.exec_card.content, text="", bg=Colors.BG_WHITE,
                                      fg=Colors.TEXT_SECONDARY, font=("Meiryo UI", 9))
        self.progress_label.pack(pady=(12, 0))
        
        self.progress_bar = ttk.Progressbar(self.exec_card.content, mode='indeterminate', length=400)
        self.progress_bar.pack(pady=(8, 0))
        
        # 認証前は無効化
        self._set_cards_enabled(False)
    
    def _set_cards_enabled(self, enabled: bool):
        """認証後のカードの有効/無効を切り替え"""
        cards = [self.brand_card, self.tag_card, self.column_card, 
                 self.date_card, self.option_card, self.preview_card, self.exec_card]
        
        for card in cards:
            card.set_disabled(not enabled)
        
        # エクスポートボタンの見た目も調整
        if enabled:
            self.export_btn.configure(bg=Colors.PRIMARY, cursor='hand2')
        else:
            self.export_btn.configure(bg='#cccccc', cursor='')
    
    def _update_preview(self):
        """出力プレビューを更新"""
        # プレビューフレームをクリア
        for widget in self.preview_frame.winfo_children():
            widget.destroy()
        
        if not self.brands:
            tk.Label(self.preview_frame, text="接続後にプレビューが表示されます",
                    bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY, 
                    font=("Meiryo UI", 9)).pack(anchor=tk.W)
            return
        
        # 選択されたブランド
        selected_brands = [self.brands[bid] for bid, var in self.brand_vars.items() if var.get()]
        
        # 選択されたカラム
        selected_columns = [col_name for col_id, col_name in AVAILABLE_COLUMNS 
                           if self.column_vars.get(col_id, tk.BooleanVar()).get()]
        
        if not selected_brands:
            tk.Label(self.preview_frame, text="ブランドを選択してください",
                    bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                    font=("Meiryo UI", 9)).pack(anchor=tk.W)
            return
        
        # シート一覧を作成
        sheets = []
        for brand in selected_brands:
            sheets.append(("data", brand, selected_columns))
        if self.ai_agent_separate.get():
            for brand in selected_brands:
                sheets.append(("data", f"{brand}_AI", selected_columns))
        
        # 統計シートのヘッダー
        stats_headers = ["日付"]
        for brand in selected_brands:
            stats_headers.extend([f"{brand}_UU", f"{brand}_回答数"])
        stats_headers.extend(["合計_UU", "合計_回答数"])
        sheets.append(("stats", "統計", stats_headers))
        
        # シート数
        tk.Label(self.preview_frame, text=f"📊 出力シート数: {len(sheets)}",
                bg=Colors.BG_WHITE, fg=Colors.TEXT,
                font=("Meiryo UI", 10, "bold")).pack(anchor=tk.W, pady=(0, 12))
        
        # 各シートのプレビュー
        for sheet_type, sheet_name, headers in sheets:
            self._draw_sheet_preview(sheet_name, headers, sheet_type)
    
    def _draw_sheet_preview(self, sheet_name: str, headers: List[str], sheet_type: str):
        """シートのテーブルプレビューを描画"""
        # シートコンテナ
        sheet_container = tk.Frame(self.preview_frame, bg=Colors.BG_WHITE)
        sheet_container.pack(fill=tk.X, pady=(0, 16))
        
        # シートタブ風のラベル
        tab_frame = tk.Frame(sheet_container, bg=Colors.BG_WHITE)
        tab_frame.pack(anchor=tk.W)
        
        tab = tk.Label(tab_frame, text=f" {sheet_name} ", 
                      bg="#4472C4", fg="white",
                      font=("Meiryo UI", 9), padx=12, pady=4)
        tab.pack(side=tk.LEFT)
        
        # 横スクロール用Canvas
        canvas_frame = tk.Frame(sheet_container, bg=Colors.BG_WHITE)
        canvas_frame.pack(fill=tk.X)
        
        canvas = tk.Canvas(canvas_frame, bg=Colors.BG_WHITE, highlightthickness=0, height=80)
        scrollbar = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=canvas.xview)
        
        canvas.configure(xscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        canvas.pack(side=tk.TOP, fill=tk.X)
        
        # テーブル本体
        table_frame = tk.Frame(canvas, bg="#999999")
        canvas.create_window((0, 0), window=table_frame, anchor="nw")
        
        # ヘッダー行
        header_row = tk.Frame(table_frame, bg="#999999")
        header_row.pack(fill=tk.X)
        
        for i, col in enumerate(headers):
            cell = tk.Label(header_row, text=col, 
                           bg="#4472C4", fg="white",
                           font=("Meiryo UI", 8, "bold"),
                           width=12, anchor="w", padx=4, pady=3)
            cell.pack(side=tk.LEFT, padx=(0, 1), pady=(0, 1))
        
        # データ行（空のダミー行を2行表示）
        for row_idx in range(2):
            data_row = tk.Frame(table_frame, bg="#999999")
            data_row.pack(fill=tk.X)
            
            row_bg = "#ffffff" if row_idx % 2 == 0 else "#f5f5f5"
            
            for i, col in enumerate(headers):
                cell = tk.Label(data_row, text="",
                               bg=row_bg, fg=Colors.TEXT,
                               font=("Meiryo UI", 8),
                               width=12, anchor="w", padx=4, pady=3)
                cell.pack(side=tk.LEFT, padx=(0, 1), pady=(0, 1))
        
        # スクロール領域を更新
        table_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))
        
        # マウスホイールで横スクロール
        def on_mouse_wheel(event):
            canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Shift-MouseWheel>", on_mouse_wheel)
        canvas.bind("<MouseWheel>", on_mouse_wheel)
    
    def _load_saved_credentials(self):
        creds = self.cred_manager.load_credentials()
        if creds:
            self.subdomain_entry.insert(0, creds.get('subdomain', ''))
            self.email_entry.insert(0, creds.get('email', ''))
            self.token_entry.insert(0, creds.get('api_token', ''))
            self.status_label.config(text="💾 認証情報を読み込みました。接続ボタンを押してください", 
                                    fg=Colors.TEXT_SECONDARY, font=("Meiryo UI", 9))
    
    def _save_credentials(self):
        subdomain = self.subdomain_entry.get().strip()
        email = self.email_entry.get().strip()
        token = self.token_entry.get().strip()
        
        if not all([subdomain, email, token]):
            messagebox.showwarning("警告", "全ての認証情報を入力してください")
            return
        
        self.cred_manager.save_credentials(subdomain, email, token)
        messagebox.showinfo("完了", "認証情報を保存しました")
    
    def _browse_output_folder(self):
        """出力先フォルダを選択（ダイアログ内から呼ばれる）"""
        folder = filedialog.askdirectory(title="出力先フォルダを選択")
        if folder:
            self._dialog_output_var.set(folder)
    
    def _show_auto_config_dialog(self):
        """自動実行設定保存ダイアログを表示"""
        if not self.api:
            messagebox.showwarning("警告", "先にZendeskに接続してください")
            return
        
        # 選択チェック
        selected_brands = [bid for bid, var in self.brand_vars.items() if var.get()]
        if not selected_brands:
            messagebox.showwarning("警告", "少なくとも1つのブランドを選択してください")
            return
        
        selected_columns = [col_id for col_id, var in self.column_vars.items() if var.get()]
        if not selected_columns:
            messagebox.showwarning("警告", "少なくとも1つのカラムを選択してください")
            return
        
        # モーダルダイアログ作成
        dialog = tk.Toplevel(self.root)
        dialog.title("自動実行用設定を保存")
        dialog.geometry("500x200")
        dialog.configure(bg=Colors.BG_WHITE)
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # 中央に配置
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 500) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 200) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # コンテンツ
        content = tk.Frame(dialog, bg=Colors.BG_WHITE, padx=24, pady=20)
        content.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(content, text="自動実行用設定を保存", bg=Colors.BG_WHITE, fg=Colors.TEXT,
                font=("Meiryo UI", 12, "bold")).pack(anchor=tk.W)
        
        tk.Label(content, text="現在のUI設定を保存します。期間は実行時に「先週月〜日」が自動設定されます。",
                bg=Colors.BG_WHITE, fg=Colors.TEXT_SECONDARY,
                font=("Meiryo UI", 9)).pack(anchor=tk.W, pady=(4, 16))
        
        # 出力先フォルダ
        folder_frame = tk.Frame(content, bg=Colors.BG_WHITE)
        folder_frame.pack(fill=tk.X, pady=(0, 16))
        
        tk.Label(folder_frame, text="出力先フォルダ:", bg=Colors.BG_WHITE,
                fg=Colors.TEXT, font=("Meiryo UI", 10)).pack(anchor=tk.W)
        
        input_frame = tk.Frame(folder_frame, bg=Colors.BG_WHITE)
        input_frame.pack(fill=tk.X, pady=(4, 0))
        
        self._dialog_output_var = tk.StringVar()
        output_entry = tk.Entry(input_frame, textvariable=self._dialog_output_var,
                               font=("Meiryo UI", 10), relief="solid", bd=1)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)
        
        browse_btn = tk.Button(input_frame, text="参照", command=self._browse_output_folder,
                              bg=Colors.BG, fg=Colors.TEXT, font=("Meiryo UI", 9),
                              relief="solid", bd=1, padx=12, cursor="hand2")
        browse_btn.pack(side=tk.LEFT, padx=(8, 0))
        
        # ボタン
        btn_frame = tk.Frame(content, bg=Colors.BG_WHITE)
        btn_frame.pack(fill=tk.X)
        
        def on_save():
            output_folder = self._dialog_output_var.get().strip()
            if not output_folder:
                messagebox.showwarning("警告", "出力先フォルダを指定してください", parent=dialog)
                return
            if not os.path.isdir(output_folder):
                messagebox.showwarning("警告", "指定されたフォルダが存在しません", parent=dialog)
                return
            
            self._save_auto_config(output_folder)
            dialog.destroy()
        
        save_btn = tk.Button(btn_frame, text="保存", command=on_save,
                            bg=Colors.PRIMARY, fg="white", font=("Meiryo UI", 10, "bold"),
                            relief="flat", padx=24, pady=6, cursor="hand2")
        save_btn.pack(side=tk.RIGHT)
        
        cancel_btn = tk.Button(btn_frame, text="キャンセル", command=dialog.destroy,
                              bg=Colors.BG, fg=Colors.TEXT, font=("Meiryo UI", 10),
                              relief="solid", bd=1, padx=16, pady=6, cursor="hand2")
        cancel_btn.pack(side=tk.RIGHT, padx=(0, 8))
    
    def _save_auto_config(self, output_folder: str):
        """自動実行用設定を保存"""
        # 選択ブランド
        selected_brands = [bid for bid, var in self.brand_vars.items() if var.get()]
        
        # 選択カラム
        selected_columns = [col_id for col_id, var in self.column_vars.items() if var.get()]
        
        # 設定データ作成（認証情報は暗号化ファイルを使うので含めない）
        config = {
            "selected_brand_ids": selected_brands,
            "exclude_tags": self.tag_selector.get_selected_tags(),
            "selected_columns": selected_columns,
            "ai_agent_separate": self.ai_agent_separate.get(),
            "output_folder": output_folder
        }
        
        # exeと同じディレクトリに保存
        config_path = os.path.join(APP_DIR, "zendesk_auto_config.json")
        
        try:
            # 認証情報も保存しておく（自動実行時に必要）
            self.cred_manager.save_credentials(
                self.subdomain_entry.get().strip(),
                self.email_entry.get().strip(),
                self.token_entry.get().strip()
            )
            
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("完了", f"自動実行設定を保存しました。\n{config_path}")
        except Exception as e:
            messagebox.showerror("エラー", f"保存に失敗しました: {e}")
    
    def _connect(self):
        subdomain = self.subdomain_entry.get().strip()
        email = self.email_entry.get().strip()
        token = self.token_entry.get().strip()
        
        if not all([subdomain, email, token]):
            messagebox.showwarning("警告", "全ての認証情報を入力してください")
            return
        
        self.status_label.config(text="接続中...", fg=Colors.TEXT_SECONDARY)
        self.root.update()
        
        try:
            self.api = ZendeskAPI(subdomain, email, token)
            
            if not self.api.test_connection():
                self.status_label.config(text="✕ 接続失敗", fg=Colors.ERROR)
                messagebox.showerror("エラー", "接続に失敗しました。認証情報を確認してください。")
                return
            
            self.status_label.config(text="ブランド取得中...", fg=Colors.TEXT_SECONDARY)
            self.root.update()
            self.brands = self.api.get_brands()
            self.forms = self.api.get_ticket_forms()
            
            self.status_label.config(text="タグ取得中...", fg=Colors.TEXT_SECONDARY)
            self.root.update()
            self.tags = self.api.get_tags()
            
            # ブランド表示更新
            for widget in self.brand_frame.winfo_children():
                widget.destroy()
            
            self.brand_vars.clear()
            for brand_id, brand_name in self.brands.items():
                var = tk.BooleanVar(value=True)
                self.brand_vars[brand_id] = var
                cb = tk.Checkbutton(self.brand_frame, text=brand_name, variable=var,
                                   bg=Colors.BG_WHITE, fg=Colors.TEXT, font=("Meiryo UI", 10),
                                   activebackground=Colors.BG_WHITE, selectcolor=Colors.BG_WHITE,
                                   command=self._update_preview)
                cb.pack(anchor=tk.W)
            
            self.tag_selector.set_tags(self.tags)
            
            # プレビュー更新
            self._update_preview()
            
            self.status_label.config(
                text=f"✓ 接続成功！ ブランド: {len(self.brands)}件, タグ: {len(self.tags)}件",
                fg=Colors.SUCCESS
            )
            
            # カードを有効化
            self._set_cards_enabled(True)
            
        except Exception as e:
            self.status_label.config(text="✕ 接続エラー", fg=Colors.ERROR)
            messagebox.showerror("エラー", f"接続エラー: {str(e)}")
    
    def _execute_export(self):
        if not self.api:
            messagebox.showwarning("警告", "先にZendeskに接続してください")
            return
        
        selected_brands = [bid for bid, var in self.brand_vars.items() if var.get()]
        if not selected_brands:
            messagebox.showwarning("警告", "少なくとも1つのブランドを選択してください")
            return
        
        selected_columns = [col_id for col_id, var in self.column_vars.items() if var.get()]
        if not selected_columns:
            messagebox.showwarning("警告", "少なくとも1つのカラムを選択してください")
            return
        
        exclude_tags = self.tag_selector.get_selected_tags()
        
        try:
            start_date_str = self.start_date_entry.get().strip()
            end_date_str = self.end_date_entry.get().strip()
            jst = pytz.timezone('Asia/Tokyo')
            start_date = jst.localize(datetime.strptime(start_date_str, "%Y-%m-%d"))
            end_date = jst.localize(datetime.strptime(end_date_str, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59))
        except ValueError:
            messagebox.showerror("エラー", "日付形式が不正です（YYYY-MM-DD）")
            return
        
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"zendesk_export_{start_date_str}_to_{end_date_str}.xlsx"
        )
        
        if not output_path:
            return
        
        def run_export():
            try:
                self.progress_bar.start()
                
                def progress_callback(msg):
                    self.progress_label.config(text=msg)
                    self.root.update()
                
                execute_export(
                    api=self.api,
                    brands=self.brands,
                    forms=self.forms,
                    selected_brand_ids=selected_brands,
                    selected_columns=selected_columns,
                    exclude_tags=exclude_tags,
                    start_date=start_date,
                    end_date=end_date,
                    ai_agent_separate=self.ai_agent_separate.get(),
                    output_path=output_path,
                    progress_callback=progress_callback
                )
                
                self.progress_bar.stop()
                self.progress_label.config(text="✓ 完了！")
                messagebox.showinfo("完了", f"エクスポートが完了しました。\n{output_path}")
                
            except Exception as e:
                self.progress_bar.stop()
                self.progress_label.config(text="✕ エラー発生")
                logger.error(f"エクスポートエラー: {e}", exc_info=True)
                messagebox.showerror("エラー", f"エクスポートエラー: {str(e)}")
        
        thread = threading.Thread(target=run_export)
        thread.daemon = True
        thread.start()
    
    def run(self):
        self.root.mainloop()


# =============================================================================
# 自動実行
# =============================================================================

def run_auto_export():
    """設定ファイルに基づいて自動エクスポートを実行"""
    config_path = os.path.join(APP_DIR, "zendesk_auto_config.json")
    
    # 設定ファイル読み込み
    if not os.path.exists(config_path):
        print(f"エラー: 設定ファイルが見つかりません: {config_path}")
        sys.exit(1)
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception as e:
        print(f"エラー: 設定ファイルの読み込みに失敗しました: {e}")
        sys.exit(1)
    
    # 認証情報読み込み
    cred_manager = CredentialManager()
    creds = cred_manager.load_credentials()
    if not creds:
        print("エラー: 認証情報が見つかりません。GUIから認証情報を保存してください。")
        sys.exit(1)
    
    # API接続
    print("Zendeskに接続中...")
    api = ZendeskAPI(creds['subdomain'], creds['email'], creds['api_token'])
    
    if not api.test_connection():
        print("エラー: Zendeskへの接続に失敗しました")
        sys.exit(1)
    
    print("接続成功")
    
    # 期間計算（先週月曜〜日曜）
    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday() + 7)
    end_of_week = start_of_week + timedelta(days=6)
    
    jst = pytz.timezone('Asia/Tokyo')
    start_date = jst.localize(start_of_week.replace(hour=0, minute=0, second=0, microsecond=0))
    end_date = jst.localize(end_of_week.replace(hour=23, minute=59, second=59, microsecond=0))
    
    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")
    
    print(f"期間: {start_date_str} 〜 {end_date_str}")
    
    # 出力パス
    output_folder = config.get('output_folder', APP_DIR)
    output_path = os.path.join(output_folder, f"zendesk_export_{start_date_str}_to_{end_date_str}.xlsx")
    
    # ブランド・フォーム取得
    print("ブランド・フォーム情報を取得中...")
    brands = api.get_brands()
    forms = api.get_ticket_forms()
    
    # 設定読み込み
    selected_brand_ids = config.get('selected_brand_ids', [])
    exclude_tags = config.get('exclude_tags', [])
    selected_columns = config.get('selected_columns', [col_id for col_id, _ in AVAILABLE_COLUMNS])
    ai_agent_separate = config.get('ai_agent_separate', True)
    
    print(f"ブランド: {[brands.get(bid, bid) for bid in selected_brand_ids]}")
    print(f"除外タグ: {exclude_tags}")
    print(f"AIエージェント別シート: {ai_agent_separate}")
    
    # 共通のエクスポート処理を実行
    execute_export(
        api=api,
        brands=brands,
        forms=forms,
        selected_brand_ids=selected_brand_ids,
        selected_columns=selected_columns,
        exclude_tags=exclude_tags,
        start_date=start_date,
        end_date=end_date,
        ai_agent_separate=ai_agent_separate,
        output_path=output_path,
        progress_callback=print
    )
    
    print(f"\n完了: {output_path}")


# =============================================================================
# メイン
# =============================================================================

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--auto":
        run_auto_export()
    else:
        app = SettingsUI()
        app.run()